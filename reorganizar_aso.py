import pandas as pd 
import numpy as np 
import openpyxl
from asig_cod_fase import tf_idf_assign
from datetime import datetime
import re
from datetime import datetime
import openpyxl
from openpyxl.styles import Color
import io

def limpiar_celdas_blancas(file_path):
    wb = openpyxl.load_workbook(file_path)
    
    COLOR_BLANCO = 'FFFFFFFF'

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Procesando hoja: {sheet_name}")
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.font and cell.font.color:
                    color = cell.font.color
                    
                    if color.type == 'rgb' and color.rgb == COLOR_BLANCO:
                        cell.value = None
                    elif color.type == 'theme' and color.theme == 0: 
                        cell.value = None

    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)  # Volver al inicio del "archivo" para poder leerlo
    return virtual_file

def aso_f(camino, anno_actual=False):


    file_path = camino
    xls = pd.ExcelFile(file_path)

    # Get all sheet names
    sheet_names = xls.sheet_names
    print(sheet_names)
    print(len(sheet_names), 'numero de hojas')

    wb = openpyxl.load_workbook(file_path, data_only=True)
    df_final = pd.DataFrame()
    for i in range(1,len(sheet_names)):
    #     print(sheet_names[i])
    # for i in range(1,4):
        print(sheet_names[i])
        ws = wb[sheet_names[i]]

        df = pd.read_excel(
            file_path,
            sheet_name=sheet_names[i],
            header=None,
            engine="openpyxl",
            keep_default_na=True,
            na_values=["", "#N/A", " ", 0]
        )

        columnas_correctas = ["SEGMENTO", "MARCA", "TIPO", "CATALOGO"]
        nombres_nuevos = df.iloc[1].fillna("").tolist()
        nuevos_nombres = columnas_correctas + nombres_nuevos[len(columnas_correctas):]
        df.columns = nuevos_nombres
        
        df = df.iloc[3:].reset_index(drop=True)
        df = df.loc[:, ~df.columns.astype(str).str.contains('%')]
        df.replace("", pd.NA, inplace=True)

        columnas_años = [col for col in df.columns if isinstance(col, (int, float))]
        columnas_2025 = [col for col in columnas_años if isinstance(col, (int, float)) and float(col) == 2025.0]
        if not columnas_2025:
            print(f"No se encontró la columna para el año 2025 en la hoja '{sheet_names[i]}'. Columnas disponibles: {columnas_años}")
            continue  # Puedes cambiar esto por raise si prefieres detener todo
        else:
            columna_nuevo = columnas_2025[0]
        columna_nuevo = columnas_2025[0]
        columnas_usado = columnas_años.copy()
        columnas_usado.remove(columna_nuevo)

        # Melt para autos nuevos
        df_nuevo = df.melt(
            id_vars=["MARCA", "CATALOGO"],
            value_vars=[columna_nuevo],
            var_name="Modelo",
            value_name="Precio"
        )
        df_nuevo["Estado"] = "Nuevo"

        # Melt para autos usados
        df_usado = df.melt(
            id_vars=["MARCA", "CATALOGO"],
            value_vars=columnas_usado,
            var_name="Modelo",
            value_name="Precio"
        )
        df_usado["Estado"] = "Usado"

        df_long = pd.concat([df_nuevo, df_usado], ignore_index=True)
        print('----------------------------')
        print(df_long.columns)
        # Convertir precios
        df_long["Precio"] = pd.to_numeric(df_long["Precio"], errors="coerce")
        df_long["Precio"] = (df_long["Precio"] * 1e6).round(-5)
        #  eliminar las filas con nan or null en precio 
        df_long = df_long.dropna(subset=["Precio"])

        # Renombrar columnas
        df_long.rename(columns={"MARCA": "Marca", "CATALOGO": "Referencia"}, inplace=True)

        if anno_actual:
            # --- ELIMINAR DUPLICADOS SOLO DE MODELO 2025 ---
            mask_2025 = df_long["Modelo"] == 2026
            df_long_2025 = df_long[mask_2025].drop_duplicates(subset=["Marca", "Referencia", "Modelo", "Estado"])
            df_long_otros = df_long[~mask_2025]

            df_long = pd.concat([df_long_2025, df_long_otros], ignore_index=True)  

        df_final = pd.concat([df_final, df_long], ignore_index=True)

    df_final['Referencia'] = df_final['Referencia'].astype(str)
    df_final = tf_idf_assign.marca_cofc(df_final)
    df_final['Servicio'] = 'Particular'
    df_final.rename(columns={"Precio": "Pricing", 'Estado' : 'ESTADO_VENTA', 'Marca' : 'MARCA', 'Modelo' : 'ANIO_MODELO', 'cod_fasecolda':'COD_FASECOLDA'}, inplace=True)
    df_final[df_final['Pricing'] !=0]
    df_final.rename(columns={"Pricing": "PRECIO_VENTA"}, inplace=True)
    df_final['COMPANIA'] = 'Asousados'
    df_final['KILOMETRAJE'] = 0
    
    return df_final



def extraer_fecha_archivo(nombre_archivo):
    meses = {
        "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
        "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
        "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12"
    }

    patron = r"(\w+)\s+(\d{1,2})\s+de\s+(\d{4})"
    match = re.search(patron, nombre_archivo, re.IGNORECASE)

    if match:
        nombre_mes = match.group(1).lower()
        dia = int(match.group(2).zfill(2))
        anio = int(match.group(3))
        
        mes_num = int(meses.get(nombre_mes))
        
        if mes_num:
            # return f"{dia}/{mes_num}/{anio}"
            list_f = [dia, mes_num, anio] 
            return list_f
    
    return "Fecha no encontrada"


# # --- Pruebas ---
# archivos = [
#     "Depreciaciones asousados estimada Diciembre 15 de 2025.xlsx",
#     "Depreciaciones asousados estimada Enero 20 de 2026.xlsx",
#     "Reporte de prueba Febrero 5 de 2024.xlsx"
# ]

# for a in archivos:
#     print(f"Original: {a} -> Resultado: {extraer_fecha_archivo(a)}")

if __name__ == '__main__':

    str_path = r'Depreciaciones asousados estimada Diciembre 15 de 2025.xlsx'
    # str_clean = r'Asousados_15_12_2025_clean.xlsx'
    str_clean = limpiar_celdas_blancas(str_path)
    df_final = aso_f(str_clean, anno_actual=False)
    l_fechas = extraer_fecha_archivo(str_path)
    # ........................................................................................
    # Faltaria automatizar la fecha, que venga del nombre que comparten el archivo
    # ........................................................................................
    try:
        fecha_dt = datetime(l_fechas[2], l_fechas[1], l_fechas[0])
    except:
        print('No se pudo obtener la fecha por el nombre, por lo que se va a poner la fecha de hoy ')
        fecha_dt = datetime.today()
    # -------------------------------------------------------------------------------------------
    # print(fecha_dt.date())
    df_final['FECHA_VENTA'] = fecha_dt
    # df_final.to_csv(f'aso_{fecha_dt}.csv', index=False)
    df_final.to_excel(f'../aso_{str(fecha_dt.date())}_clean_try2.xlsx', index=False)