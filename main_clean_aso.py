import openpyxl
from openpyxl.styles import Color

def limpiar_celdas_blancas_server(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    
    COLOR_BLANCO = 'FFFFFFFF'

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"Procesando hoja: {sheet_name}")
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.font and cell.font.color:
                    color = cell.font.color
                    
                    if color.type == 'rgb' and color.rgb == COLOR_BLANCO:
                        cell.value = None
                    elif color.type == 'theme' and color.theme == 0: 
                        cell.value = None

    wb.save(output_path)
    print(f"Archivo guardado en: {output_path}")

limpiar_celdas_blancas_server('Asousados_15_12_2025.xlsx', 'Asousados_15_12_2025_clean.xlsx')